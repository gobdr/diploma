import json
import os
import pathlib
from collections import defaultdict

import pandas as pd
from matplotlib import pyplot as plt
from openpyxl import load_workbook


pd.set_option('display.max_rows', None)
pd.set_option('display.max_columns', None)
pd.set_option('display.max_colwidth', None)
pd.set_option('display.width', 10000)

class ExcelToFormattedCSV:
    def __init__(self, file_path):
        self.file_path = file_path
        self.wb = load_workbook(filename=file_path)
        self.sheets_to_ignore = ['Intro']

    def _group_columns_by_prefix(self, df):
        # Группировка столбцов по префиксам
        groups = defaultdict(list)
        for col in df.columns:
            if pd.notnull(col) and col != 'Category'and col !=  'Subcategory':
                groups[col].append(col)
        return groups

    def _format_row(self, row, column_groups):
        # print(row)
        line_parts = [str(row['Category']) if pd.notnull(row['Category']) else "", str(row['Subcategory']) if pd.notnull(row['Subcategory']) else ""]
        for _, cols in column_groups.items():
            if 'x(' not in str(cols[0]) :
                if len(cols) == 2:  # Если есть пара столбцов для диапазона
                    # print(cols)
                    # print(row[cols[0]][0], row[cols[1]][1])
                    lower_value = row[cols[0]][0] if pd.notnull(row[cols[0]][0]) else ''
                    upper_value = row[cols[1]][1] if pd.notnull(row[cols[1]][1]) else ''
                    if '' == lower_value or '' == upper_value:
                        line_parts = []
                        break
                    formatted_value = f"[{lower_value};{upper_value}]"
                else:  # Если столбец один
                    print(cols, len(cols))
                    # print(row[cols][])
                    formatted_value = f"{row[cols[0]][0]}" if pd.notnull(row[cols[0]][0]) else ""
                line_parts.append(formatted_value)
        return '|'.join(line_parts).replace('\n','')

    def _format_data(self, df):
        print(df.head(3))

        formatted_lines = ["Category|Subcategory|Delta|Atomic weight"]

        df.iloc[0, 0] = 'Category'
        df.iloc[0, 1] = 'Subcategory'
        df.columns = df.iloc[0]
        df = df.drop(df.index[0])
        df = df.reset_index(drop=True)

        print(df.head(3))

        column_groups = self._group_columns_by_prefix(df)
        print(column_groups)
        for _, row in df.iterrows():
            formatted_lines.append(self._format_row(row, column_groups))
        return formatted_lines

    def process_sheets(self):
        for sheet_name in self.wb.sheetnames:
            print(sheet_name)
            if sheet_name in self.sheets_to_ignore:
                continue
            df = pd.read_excel(self.file_path, sheet_name=sheet_name)
            df.dropna(how='all', inplace=True)  # Удаление пустых строк
            formatted_data = self._format_data(df)
            output_file_path = str(pathlib.Path(pathlib.Path.cwd().parent, 'data', 'dataset3', 'lists', f'{sheet_name}.csv'))
            with open(output_file_path, 'w', encoding='utf-8') as file:
                for line in formatted_data:
                    file.write(line + "\n")
            print(f"Processed {sheet_name}, saved to {output_file_path}")
    def save_intervals_fig(self):
        def parse_range(value):
            lower, upper = value.strip('[]').split(';')
            lower = float(lower.replace(',', '.'))
            upper = float(upper.replace(',', '.'))
            return lower, upper

        for sheet_name in self.wb.sheetnames:
            print(sheet_name)
            if sheet_name in self.sheets_to_ignore:
                continue
            workdir = str(pathlib.Path(pathlib.Path.cwd().parent, 'data', 'dataset3', 'lists' , f'{sheet_name}.csv'))
            df = pd.read_csv(workdir, sep='|')
            df['Subcategory'] = df['Subcategory'].fillna(df['Category'])

            category_data = df #[df['Category'] == category]

            # Extract lower, upper bounds and calculate the means for the current category plot
            category_lower_bounds = []
            category_upper_bounds = []
            category_means = []
            for value in category_data['Delta']:
                lower, upper = parse_range(value)
                category_lower_bounds.append(lower)
                category_upper_bounds.append(upper)
                category_means.append((upper + lower) / 2)

            # Calculate errors from the mean to the lower and upper bounds for current category data
            category_errors_lower = [mean - lower for mean, lower in zip(category_means, category_lower_bounds)]
            category_errors_upper = [upper - mean for mean, upper in zip(category_means, category_upper_bounds)]

            # Subcategory numbers for the y-axis for current category data
            category_subcategory_numbers = list(range(1, len(category_means) + 1))

            # Create the plot for the current category
            fig, ax = plt.subplots(figsize=(10, 6))

            # Plot horizontal bars
            ax.errorbar(category_means, category_subcategory_numbers, xerr=[category_errors_lower, category_errors_upper], linestyle='None', ecolor='blue', capsize=5)

            # Set labels
            ax.set_xlabel('δ, x 10³')
            # ax.set_ylabel('Subcategory')

            # Set the title to the category name
            ax.set_title(f'Variations  in {sheet_name}')

            # Add subcategory names as y-tick labels without the category name
            ax.set_yticks(category_subcategory_numbers)
            ax.set_yticklabels(category_data['Subcategory'], rotation='horizontal')

            # Invert the y-axis to match the given image
            ax.invert_yaxis()

            # Show the plot with a tight layout to ensure the labels fit well
            plt.tight_layout()
            os.makedirs(str(pathlib.Path(pathlib.Path.cwd(), 'doc', 'out', f'{sheet_name}')), exist_ok=True)
            plt.savefig(str(pathlib.Path(pathlib.Path.cwd(), 'doc', 'out', f'{sheet_name}', f'{sheet_name}_intervals.png')))
            # plt.show()

    def save_mode_fig(self):

        def merge_intervals(intervals):
            if not intervals:
                return []

            # Сортировка интервалов по начальной точке
            intervals.sort(key=lambda x: x[0])

            merged = [intervals[0]]  # Добавляем первый интервал в результат
            for current in intervals[1:]:
                prev = merged[-1]  # Последний добавленный в результат интервал

                # Если текущий интервал пересекается с последним объединённым интервалом
                if current[0] <= prev[1]:
                    # Объединяем, обновляя конечную точку предыдущего интервала, если необходимо
                    merged[-1][1] = max(prev[1], current[1])
                else:
                    # Если интервалы не пересекаются, добавляем текущий интервал к результату
                    merged.append(current)

            return merged
        for sheet_name in self.wb.sheetnames:
            print(sheet_name)
            if sheet_name in self.sheets_to_ignore:
                continue
            workdir = str(pathlib.Path(pathlib.Path.cwd().parent, 'data', 'dataset3', 'lists', f'{sheet_name}.csv'))
            data = pd.read_csv(workdir, sep='|')
            data['Interval'] = data['Delta'].apply(lambda x: tuple(map(float, x.strip('[]').split(';'))))

            # Функция для расчета моды на каждом участке
            def calculate_mode(intervals):
                points = sorted(set([point for interval in intervals for point in interval]))
                segment_modes = []
                for i in range(len(points) - 1):
                    left, right = points[i], points[i + 1]
                    mode_count = sum(left >= start and right <= end for start, end in intervals)
                    segment_modes.append((left, right, mode_count))
                return segment_modes

            # Вычисление моды для каждого сегмента
            segment_modes = calculate_mode(data['Interval'].tolist())

            # Построение графика
            plt.figure(figsize=(12, 6))
            for left, right, mode in segment_modes:
                plt.plot([left, right], [mode, mode], '-o', color='blue', lw=2, markersize=4)

            # Нахождение и выделение сегмента с максимальной модой
            max_mode_value = max(mode for _, _, mode in segment_modes)
            max_mode_segments = [[left, right] for left, right, mode in segment_modes if mode == max_mode_value]
            max_mode_segments = merge_intervals(max_mode_segments)
            for left, right in max_mode_segments:
                plt.plot([left, right], [max_mode_value, max_mode_value], '-o', color='red', lw=2, markersize=4)

            # Добавление вертикальных пунктирных линий
            # for left, _, mode in segment_modes[:-1]:
            #     plt.vlines(left, ymin=0, ymax=max_mode_value, color='blue', linestyles='dashed', lw=1)

            # Добавление аннотаций к сегментам
            for left, right, mode in segment_modes:
                plt.text((left + right) / 2, mode, f"{mode}", ha='center', va='bottom')

            plt.xlabel('x')
            plt.ylabel('Мода μi')
            plt.title('Интервалы и их моды')
            plt.grid(True)
            plt.ylim(0, max_mode_value + 1)
            os.makedirs(str(pathlib.Path(pathlib.Path.cwd(), 'doc', 'out', f'{sheet_name}')), exist_ok=True)
            plt.savefig(str(pathlib.Path(pathlib.Path.cwd(), 'doc', 'out', f'{sheet_name}', f'{sheet_name}_mode.png')))
            # plt.show()
            data_dict = {
                "max_mode": max_mode_value,
                "max_iterval": max_mode_segments
            }
            with open(str(pathlib.Path(pathlib.Path.cwd(), 'doc', 'out', f'{sheet_name}', f'{sheet_name}_mode.json')), "w") as json_file:
                json.dump(data_dict, json_file, indent=4)

            # Вывод максимальной моды и соответствующего интервала
            print(f"Максимальная мода: {max_mode_value}")
            print(f"Интервал с максимальной модой: {max_mode_segments}")



    def calculate_parameters(self):
        def parse_interval(interval_str):
            print(interval_str, type(interval_str))
            interval_str = str(interval_str).replace('[', '').replace(']', '')
            try:
                lower, upper = interval_str.split(';')
            except Exception as e:
                try:

                    lower, upper = interval_str.split(',')
                except Exception as e:
                    return float(0), float(0)
            return float(lower), float(upper)

        def jaccard_index(interval_x, interval_y):
            lower_x, upper_x = parse_interval(list(interval_x))
            lower_y, upper_y = parse_interval(list(interval_y))

            intersection_lower = max(lower_x, lower_y)
            intersection_upper = min(upper_x, upper_y)
            wid_x =  abs(upper_x - lower_x)

            # Если интервалы не пересекаются, возвращаем 0
            if wid_x == 0:
                return 'Error'
            else:
                return round((intersection_upper - intersection_lower) / wid_x, 4)

        for sheet_name in self.wb.sheetnames:

            print(sheet_name)
            if sheet_name in self.sheets_to_ignore:
                continue
            # Parse the intervals in the dataframe
            workdir = str(pathlib.Path(pathlib.Path.cwd().parent, 'data', 'dataset3', 'lists', f'{sheet_name}.csv'))
            data = pd.read_csv(workdir, sep='|')
            data['Parsed Delta'] = data['Delta'].apply(parse_interval)

            # Extract the mode intervals from the JSON data
            with open(str(pathlib.Path(pathlib.Path.cwd(), 'doc', 'out', f'{sheet_name}', f'{sheet_name}_mode.json')), 'r') as file:
                carbon_mode = json.load(file)
            print(carbon_mode)
            max_interval = carbon_mode['max_iterval'][0]

            max_interval = (float(max_interval[0]), float(max_interval[1]))  # Convert to tuple
            print(max_interval)
            # Calculate the Jaccard index for each row in the dataframe
            # If Subcategory is NaN, use Category instead
            data['Subcategory'] = data['Subcategory'].fillna(data['Category'])
            data['JaccardIndex'] = data.apply(lambda row: jaccard_index(row['Parsed Delta'], max_interval), axis=1)
            # print(data[data['Jaccard Index']<0])
            # Select the relevant columns to save to a new file
            print(data.columns)
            data['maxInterval'] = str(max_interval)

            output_df = data[['Category', 'Subcategory', 'Delta', 'maxInterval', 'JaccardIndex']]

            os.makedirs(str(pathlib.Path(pathlib.Path.cwd(), 'doc', 'out', f'{sheet_name}')), exist_ok=True)
            output_file_path = str(pathlib.Path(pathlib.Path.cwd(), 'doc', 'out', f'{sheet_name}', f'{sheet_name}_Jaccard_Index.csv'))
            output_df.to_csv(output_file_path, index=False)

            output_file_path, output_df.head()

# Пример использования:
file_path = str(pathlib.Path(pathlib.Path.cwd().parent,'data','dataset3','Dataset Isotope.xlsx'))
converter = ExcelToFormattedCSV(file_path)
converter.save_intervals_fig()
converter.save_mode_fig()
converter.calculate_parameters()
# С
# N
# O
# S
# Hydrogen