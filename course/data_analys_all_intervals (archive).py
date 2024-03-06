import pathlib

import pandas as pd
from openpyxl import load_workbook


class ExcelToCSVConverter:
    def __init__(self, file_path):
        self.file_path = file_path
        self.wb = load_workbook(filename=file_path)

    @staticmethod
    def format_value(lower, upper):
        """Форматирует значения в формате [lower;upper]."""
        return f"[{lower};{upper}]" if pd.notnull(lower) and pd.notnull(upper) else ""

    def convert_sheets(self):
        for sheet_name in self.wb.sheetnames:
            if sheet_name == 'Intro':  # Пропускаем начальный лист, если он не содержит данных
                continue
            self.process_sheet(sheet_name)

    def process_sheet(self, sheet_name):
        df = pd.read_excel(self.file_path, sheet_name=sheet_name)
        # Удаление строк, где нет данных по 'Atomic weight'
        print(df.head(3))
        try:
            df = df.dropna(subset=['Atomic weight'])
        except Exception as e:
            print(e.args[0])
        # Формируем строки с разделителями "|"
        formatted_lines = ["Category|Subcategory|103 δ34SVCDT|x(32S)|x(33S)|x(34S)|x(36S)|Atomic weight"]  # Заголовок
        for _, row in df.iterrows():
            line_parts = [str(row[col]) for col in df.columns if col not in ['Reference', 'Empty']]  # Исключение ненужных столбцов
            line = '|'.join(line_parts)
            formatted_lines.append(line.replace(";", "|"))  # Замена ";" на "|"

        # Сохранение отформатированных данных в CSV файл
        csv_file_path = str(pathlib.Path(pathlib.Path.cwd().parent, 'data', 'dataset3', 'lists' ,f'{sheet_name}.csv'))

        with open(csv_file_path, 'w', encoding='utf-8') as file:
            for line in formatted_lines:
                file.write(line + "\n")
        print(f"Файл {csv_file_path} сохранен.")


# Использование класса
file_path = str(pathlib.Path(pathlib.Path.cwd().parent,'data','dataset3','isotope_data.xlsx'))

converter = ExcelToCSVConverter(file_path)
converter.convert_sheets()
